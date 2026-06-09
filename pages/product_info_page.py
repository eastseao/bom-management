#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""成品信息页面"""

import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk


class ProductInfoPage(ctk.CTkFrame):
    def __init__(self, parent, db, colors):
        super().__init__(parent, fg_color=colors["bg"], corner_radius=0)
        self.C = colors
        self.db = db
        self._all_data = []
        self._build()

    def _build(self):
        # ── 顶部标题栏 ─────────────────────────────
        header = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=0, height=64)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header, text="📋  成品信息",
            font=ctk.CTkFont(family="Microsoft YaHei", size=20, weight="bold"),
            text_color=self.C["text"],
        ).pack(side="left", padx=24, pady=16)

        # ── 筛选区 ─────────────────────────────────
        filter_frame = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=10)
        filter_frame.pack(fill="x", padx=16, pady=(16, 8))

        filter_inner = ctk.CTkFrame(filter_frame, fg_color="transparent")
        filter_inner.pack(fill="x", padx=16, pady=12)

        # 第一行：渠道 + 项目号 + 品名
        row1 = ctk.CTkFrame(filter_inner, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 6))

        ctk.CTkLabel(row1, text="渠道", font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                     text_color=self.C["text_secondary"], width=50).pack(side="left", padx=(0, 4))

        self.channel_var = tk.StringVar(value="所有渠道")
        self.channel_combo = ctk.CTkComboBox(
            row1, values=["所有渠道", "传统渠道", "电商渠道", "B端渠道", "其他渠道"],
            variable=self.channel_var, width=130,
            font=ctk.CTkFont(family="Microsoft YaHei", size=12),
            fg_color=self.C["bg"], border_color=self.C["border"],
            button_color=self.C["primary"], button_hover_color=self.C["primary_hover"],
            dropdown_fg_color=self.C["card"], dropdown_text_color=self.C["text"],
            dropdown_hover_color=self.C["sidebar_hover"],
        )
        self.channel_combo.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row1, text="项目号", font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                     text_color=self.C["text_secondary"], width=50).pack(side="left", padx=(0, 4))

        self.project_no_entry = ctk.CTkEntry(
            row1, width=140,
            font=ctk.CTkFont(family="Microsoft YaHei", size=12),
            fg_color=self.C["bg"], border_color=self.C["border"],
        )
        self.project_no_entry.pack(side="left", padx=(0, 16))

        ctk.CTkLabel(row1, text="品名", font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                     text_color=self.C["text_secondary"], width=50).pack(side="left", padx=(0, 4))

        self.product_name_entry = ctk.CTkEntry(
            row1, width=140,
            font=ctk.CTkFont(family="Microsoft YaHei", size=12),
            fg_color=self.C["bg"], border_color=self.C["border"],
        )
        self.product_name_entry.pack(side="left")

        # 第二行：按钮
        row2 = ctk.CTkFrame(filter_inner, fg_color="transparent")
        row2.pack(fill="x")

        self.query_btn = ctk.CTkButton(
            row2, text="🔍 查询", width=90, height=34,
            fg_color=self.C["primary"], hover_color=self.C["primary_hover"],
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            command=self._on_query,
        )
        self.query_btn.pack(side="left", padx=(0, 8))

        self.reset_btn = ctk.CTkButton(
            row2, text="🔄 重置", width=90, height=34,
            fg_color="transparent", text_color=self.C["primary"],
            border_color=self.C["border"], border_width=1,
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            command=self._on_reset,
        )
        self.reset_btn.pack(side="left", padx=(0, 16))

        self.import_btn = ctk.CTkButton(
            row2, text="📥 导入Excel", width=100, height=34,
            fg_color=self.C["success"], hover_color="#7A9472",
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            command=self._on_import,
        )
        self.import_btn.pack(side="left", padx=(0, 8))

        self.export_btn = ctk.CTkButton(
            row2, text="📤 导出Excel", width=100, height=34,
            fg_color=self.C["warning"], hover_color="#B89A5D",
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            command=self._on_export,
        )
        self.export_btn.pack(side="left")

        # ── 表格区 ─────────────────────────────────
        table_frame = ctk.CTkFrame(self, fg_color=self.C["card"], corner_radius=10)
        table_frame.pack(fill="both", expand=True, padx=16, pady=(8, 16))

        # 使用 tkinter Treeview 实现表格
        self._build_treeview(table_frame)

    def _build_treeview(self, parent):
        """构建Treeview表格"""
        # 样式
        style = ttk.Style()
        style.theme_use("clam")

        bg_color = self._hex_to_rgb(self.C["card"])
        style.configure("Product.Treeview",
                        background=self.C["card"],
                        fieldbackground=self.C["card"],
                        foreground=self.C["text"],
                        rowheight=32,
                        font=("Microsoft YaHei", 11),
                        borderwidth=0)
        style.configure("Product.Treeview.Heading",
                        background=self.C["primary"],
                        foreground="#FFFFFF",
                        font=("Microsoft YaHei", 11, "bold"),
                        relief="flat",
                        borderwidth=0,
                        padding=(6, 6))
        style.map("Product.Treeview.Heading",
                  background=[("active", self.C["primary_hover"])])
        style.map("Product.Treeview",
                  background=[("selected", self.C["primary_light"])],
                  foreground=[("selected", self.C["text"])])

        # 列定义
        self.columns = ["项目号", "品名", "规格", "单位", "零售价", "产品属性", "保质期（天）", "品牌", "可供应渠道"]

        # Treeview
        tree_container = ctk.CTkFrame(parent, fg_color="transparent")
        tree_container.pack(fill="both", expand=True, padx=8, pady=8)

        self.tree = ttk.Treeview(
            tree_container,
            columns=self.columns,
            show="headings",
            style="Product.Treeview",
        )

        # 设置列标题和宽度
        col_widths = [100, 150, 80, 60, 80, 100, 100, 100, 120]
        for col, width in zip(self.columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center", minwidth=60)

        # 滚动条
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        # 双击编辑
        self.tree.bind("<Double-1>", self._on_double_click)

    def _on_query(self):
        """查询成品信息"""
        channel = self.channel_var.get()
        project_no = self.project_no_entry.get().strip()
        product_name = self.product_name_entry.get().strip()

        data = self.db.get_finished_products(
            channel=channel,
            project_no=project_no if project_no else None,
            product_name=product_name if product_name else None,
        )
        self._all_data = data
        self._refresh_table(data)

    def _on_reset(self):
        """重置筛选条件"""
        self.channel_var.set("所有渠道")
        self.project_no_entry.delete(0, "end")
        self.product_name_entry.delete(0, "end")
        self._all_data = []
        self._clear_table()

    def _on_import(self):
        """从Excel导入成品信息"""
        filepath = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active

            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                rows.append({
                    "project_no": str(row[0]) if row[0] else "",
                    "product_name": str(row[1]) if row[1] else "",
                    "spec": str(row[2]) if row[2] else "",
                    "unit": str(row[3]) if row[3] else "",
                    "retail_price": float(row[4]) if row[4] and str(row[4]).strip() else 0,
                    "product_attribute": str(row[5]) if row[5] else "",
                    "shelf_life_days": int(row[6]) if row[6] and str(row[6]).strip() else 0,
                    "brand": str(row[7]) if row[7] else "",
                    "supply_channel": str(row[8]) if row[8] else "",
                })

            if rows:
                self.db.import_finished_products(rows)
                messagebox.showinfo("导入成功", f"成功导入 {len(rows)} 条成品信息")
                self._on_query()
            else:
                messagebox.showwarning("提示", "未找到有效数据")
        except Exception as e:
            messagebox.showerror("导入失败", f"导入Excel时出错：\n{e}")

    def _on_export(self):
        """导出成品信息到Excel"""
        if not self._all_data:
            messagebox.showwarning("提示", "当前没有数据可导出，请先查询")
            return

        filepath = filedialog.asksaveasfilename(
            title="导出Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="成品信息导出.xlsx",
        )
        if not filepath:
            return

        try:
            headers = self.columns
            rows = []
            for d in self._all_data:
                rows.append({
                    "项目号": d.get("project_no", ""),
                    "品名": d.get("product_name", ""),
                    "规格": d.get("spec", ""),
                    "单位": d.get("unit", ""),
                    "零售价": d.get("retail_price", 0),
                    "产品属性": d.get("product_attribute", ""),
                    "保质期（天）": d.get("shelf_life_days", 0),
                    "品牌": d.get("brand", ""),
                    "可供应渠道": d.get("supply_channel", ""),
                })
            self.db.export_to_xlsx(filepath, "成品信息", headers, rows)
            messagebox.showinfo("导出成功", f"成功导出 {len(rows)} 条记录")
        except Exception as e:
            messagebox.showerror("导出失败", f"导出Excel时出错：\n{e}")

    def _on_double_click(self, event):
        """双击编辑行"""
        item = self.tree.selection()
        if not item:
            return
        values = self.tree.item(item[0], "values")
        idx = self.tree.index(item[0])
        if idx >= len(self._all_data):
            return

        record = self._all_data[idx]
        self._show_edit_dialog(record)

    def _show_edit_dialog(self, record):
        """显示编辑对话框"""
        dialog = tk.Toplevel(self)
        dialog.title("编辑成品信息" if record.get("id") else "新增成品信息")
        dialog.geometry("500x500")
        dialog.resizable(False, False)
        dialog.configure(bg=self.C["card"])
        dialog.transient(self)
        dialog.grab_set()

        # 居中
        dialog.update_idletasks()
        dw, dh = 500, 500
        sw = dialog.winfo_screenwidth()
        sh = dialog.winfo_screenheight()
        dialog.geometry(f"{dw}x{dh}+{(sw-dw)//2}+{(sh-dh)//2}")

        fields = [
            ("项目号", "project_no"),
            ("品名", "product_name"),
            ("规格", "spec"),
            ("单位", "unit"),
            ("零售价", "retail_price"),
            ("产品属性", "product_attribute"),
            ("保质期（天）", "shelf_life_days"),
            ("品牌", "brand"),
        ]

        entries = {}
        for i, (label, key) in enumerate(fields):
            ctk.CTkLabel(dialog, text=label,
                         font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                         text_color=self.C["text"]).pack(anchor="w", padx=24, pady=(10, 2))

            var = tk.StringVar(value=str(record.get(key, "")) if record.get(key) is not None else "")
            entry = ctk.CTkEntry(
                dialog, textvariable=var, width=440,
                font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                fg_color=self.C["bg"], border_color=self.C["border"],
            )
            entry.pack(padx=24)
            entries[key] = var

        # 渠道下拉
        ctk.CTkLabel(dialog, text="可供应渠道",
                     font=ctk.CTkFont(family="Microsoft YaHei", size=12),
                     text_color=self.C["text"]).pack(anchor="w", padx=24, pady=(10, 2))

        channel_var = tk.StringVar(value=record.get("supply_channel", "其他渠道"))
        channel_combo = ctk.CTkComboBox(
            dialog, values=["传统渠道", "电商渠道", "B端渠道", "其他渠道"],
            variable=channel_var, width=440,
            font=ctk.CTkFont(family="Microsoft YaHei", size=12),
            fg_color=self.C["bg"], border_color=self.C["border"],
            button_color=self.C["primary"], button_hover_color=self.C["primary_hover"],
            dropdown_fg_color=self.C["card"], dropdown_text_color=self.C["text"],
        )
        channel_combo.pack(padx=24)

        def _save():
            data = {key: var.get().strip() for key, var in entries.items()}
            data["supply_channel"] = channel_var.get()

            # 数值转换
            try:
                data["retail_price"] = float(data["retail_price"]) if data["retail_price"] else 0
            except ValueError:
                data["retail_price"] = 0
            try:
                data["shelf_life_days"] = int(data["shelf_life_days"]) if data["shelf_life_days"] else 0
            except ValueError:
                data["shelf_life_days"] = 0

            if not data["project_no"] or not data["product_name"]:
                messagebox.showwarning("提示", "项目号和品名不能为空")
                return

            if record.get("id"):
                self.db.update_finished_product(record["id"], data)
            else:
                self.db.save_finished_product(data)

            dialog.destroy()
            self._on_query()
            messagebox.showinfo("成功", "保存成功")

        btn_frame = ctk.CTkFrame(dialog, fg_color="transparent")
        btn_frame.pack(fill="x", padx=24, pady=(20, 16))

        ctk.CTkButton(
            btn_frame, text="✓ 保存", width=100, height=36,
            fg_color=self.C["success"], hover_color="#7A9472",
            font=ctk.CTkFont(family="Microsoft YaHei", size=13, weight="bold"),
            command=_save,
        ).pack(side="right", padx=(8, 0))

        ctk.CTkButton(
            btn_frame, text="取消", width=80, height=36,
            fg_color="transparent", text_color=self.C["text_secondary"],
            border_color=self.C["border"], border_width=1,
            font=ctk.CTkFont(family="Microsoft YaHei", size=13),
            command=dialog.destroy,
        ).pack(side="right")

    def _refresh_table(self, data):
        """刷新表格数据"""
        self._clear_table()
        for d in data:
            values = (
                d.get("project_no", ""),
                d.get("product_name", ""),
                d.get("spec", ""),
                d.get("unit", ""),
                d.get("retail_price", 0),
                d.get("product_attribute", ""),
                d.get("shelf_life_days", 0),
                d.get("brand", ""),
                d.get("supply_channel", ""),
            )
            self.tree.insert("", "end", values=values)

    def _clear_table(self):
        """清空表格"""
        for item in self.tree.get_children():
            self.tree.delete(item)

    @staticmethod
    def _hex_to_rgb(hex_color):
        """HEX颜色转RGB"""
        hex_color = hex_color.lstrip("#")
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
