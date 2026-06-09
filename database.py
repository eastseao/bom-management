#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""BOM管理系统 - 数据库模块"""

import sqlite3
import os
import json


def _get_data_dir(data_dir=None):
    if data_dir is not None:
        return data_dir
    return os.path.join(os.path.expanduser("~"), "BOM管理系统数据")


class Database:
    def __init__(self, data_dir=None):
        self.data_dir = _get_data_dir(data_dir)
        os.makedirs(self.data_dir, exist_ok=True)
        self.db_path = os.path.join(self.data_dir, "bom_system.db")
        db_uri = self.db_path.replace("\\", "/")
        try:
            self.conn = sqlite3.connect(f"file:{db_uri}?mode=rwc", uri=True,
                                        check_same_thread=False)
        except Exception:
            self.conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self.conn.execute("PRAGMA journal_mode=MEMORY")
        self.conn.row_factory = sqlite3.Row
        self._init_tables()

    def _init_tables(self):
        c = self.conn.cursor()

        # ── 成品信息表 ──
        c.execute("""
            CREATE TABLE IF NOT EXISTS finished_products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_no TEXT NOT NULL,
                product_name TEXT NOT NULL,
                spec TEXT,
                unit TEXT DEFAULT '',
                retail_price REAL DEFAULT 0,
                product_attribute TEXT,
                shelf_life_days INTEGER DEFAULT 0,
                brand TEXT,
                supply_channel TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        # ── 成品BOM表 ──
        c.execute("""
            CREATE TABLE IF NOT EXISTS product_bom (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                finished_project_no TEXT NOT NULL,
                product_name TEXT,
                spec TEXT,
                retail_price REAL DEFAULT 0,
                brand TEXT,
                material_project_no TEXT NOT NULL,
                material_name TEXT,
                quantity REAL DEFAULT 0,
                unit TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )
        """)

        self.conn.commit()

    # ====== 成品信息 CRUD ======
    def get_finished_products(self, channel=None, project_no=None, product_name=None):
        c = self.conn.cursor()
        sql = "SELECT * FROM finished_products WHERE 1=1"
        params = []
        if channel and channel != "所有渠道":
            sql += " AND supply_channel=?"
            params.append(channel)
        if project_no:
            sql += " AND project_no LIKE ?"
            params.append(f"%{project_no}%")
        if product_name:
            sql += " AND product_name LIKE ?"
            params.append(f"%{product_name}%")
        sql += " ORDER BY id DESC"
        c.execute(sql, params)
        return [dict(r) for r in c.fetchall()]

    def save_finished_product(self, data):
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO finished_products(project_no, product_name, spec, unit,
            retail_price, product_attribute, shelf_life_days, brand, supply_channel)
            VALUES(:project_no, :product_name, :spec, :unit,
            :retail_price, :product_attribute, :shelf_life_days, :brand, :supply_channel)
        """, data)
        self.conn.commit()
        return c.lastrowid

    def update_finished_product(self, pid, data):
        c = self.conn.cursor()
        c.execute("""
            UPDATE finished_products SET project_no=:project_no, product_name=:product_name,
            spec=:spec, unit=:unit, retail_price=:retail_price,
            product_attribute=:product_attribute, shelf_life_days=:shelf_life_days,
            brand=:brand, supply_channel=:supply_channel
            WHERE id=:id
        """, {**data, "id": pid})
        self.conn.commit()

    def delete_finished_product(self, pid):
        c = self.conn.cursor()
        c.execute("DELETE FROM finished_products WHERE id=?", (pid,))
        self.conn.commit()

    def _clean_value(self, val, val_type="str"):
        """清洗单个值：将 'null'/'None'/''/None 转为合理的默认值"""
        if val is None:
            return None if val_type == "none" else ("", 0.0, 0)[
                ("str", "float", "int").index(val_type)]
        s = str(val).strip()
        if s.lower() in ("null", "none", "na", "n/a", ""):
            return None if val_type == "none" else ("", 0.0, 0)[
                ("str", "float", "int").index(val_type)]
        if val_type == "float":
            try:
                return float(s)
            except (ValueError, TypeError):
                return 0.0
        if val_type == "int":
            try:
                return int(float(s))
            except (ValueError, TypeError):
                return 0
        return s

    def import_finished_products(self, rows):
        """批量导入成品信息（自动清洗数据）"""
        c = self.conn.cursor()
        for row in rows:
            cleaned = {
                "project_no": self._clean_value(row.get("项目号", row.get("project_no", "")), "str") or "",
                "product_name": self._clean_value(row.get("品名", row.get("product_name", "")), "str") or "",
                "spec": self._clean_value(row.get("规格", row.get("spec", "")), "str"),
                "unit": self._clean_value(row.get("单位", row.get("unit", "")), "str"),
                "retail_price": self._clean_value(row.get("零售价", row.get("retail_price", 0)), "float"),
                "product_attribute": self._clean_value(row.get("产品属性", row.get("product_attribute", "")), "str"),
                "shelf_life_days": self._clean_value(row.get("保质期（天）", row.get("shelf_life_days", 0)), "int"),
                "brand": self._clean_value(row.get("品牌", row.get("brand", "")), "str"),
                "supply_channel": self._clean_value(row.get("可供应渠道", row.get("supply_channel", "")), "str"),
            }
            if not cleaned["project_no"] or not cleaned["product_name"]:
                continue
            c.execute("""
                INSERT INTO finished_products(project_no, product_name, spec, unit,
                retail_price, product_attribute, shelf_life_days, brand, supply_channel)
                VALUES(:project_no, :product_name, :spec, :unit,
                :retail_price, :product_attribute, :shelf_life_days, :brand, :supply_channel)
            """, cleaned)
        self.conn.commit()

    # ====== 成品BOM CRUD ======
    def get_product_bom(self, product_name=None, finished_project_no=None,
                        material_project_no=None, material_name=None):
        c = self.conn.cursor()
        sql = "SELECT * FROM product_bom WHERE 1=1"
        params = []
        if product_name:
            sql += " AND product_name LIKE ?"
            params.append(f"%{product_name}%")
        if finished_project_no:
            sql += " AND finished_project_no LIKE ?"
            params.append(f"%{finished_project_no}%")
        if material_project_no:
            sql += " AND material_project_no LIKE ?"
            params.append(f"%{material_project_no}%")
        if material_name:
            sql += " AND material_name LIKE ?"
            params.append(f"%{material_name}%")
        sql += " ORDER BY finished_project_no, id"
        c.execute(sql, params)
        return [dict(r) for r in c.fetchall()]

    def save_product_bom(self, data):
        """保存BOM记录（自动清洗数据）"""
        cleaned = {
            "finished_project_no": self._clean_value(data.get("成品项目号", data.get("finished_project_no", "")), "str") or "",
            "product_name": self._clean_value(data.get("品名", data.get("product_name", "")), "str"),
            "spec": self._clean_value(data.get("规格", data.get("spec", "")), "str"),
            "retail_price": self._clean_value(data.get("零售价（元）", data.get("retail_price", 0)), "float"),
            "brand": self._clean_value(data.get("品牌", data.get("brand", "")), "str"),
            "material_project_no": self._clean_value(data.get("物料项目号", data.get("material_project_no", "")), "str") or "",
            "material_name": self._clean_value(data.get("物料名称", data.get("material_name", "")), "str"),
            "quantity": self._clean_value(data.get("数量", data.get("quantity", 0)), "float"),
            "unit": self._clean_value(data.get("计量单位", data.get("unit", "")), "str"),
        }
        c = self.conn.cursor()
        c.execute("""
            INSERT INTO product_bom(finished_project_no, product_name, spec,
            retail_price, brand, material_project_no, material_name, quantity, unit)
            VALUES(:finished_project_no, :product_name, :spec,
            :retail_price, :brand, :material_project_no, :material_name, :quantity, :unit)
        """, cleaned)
        self.conn.commit()
        return c.lastrowid

    def update_product_bom(self, bid, data):
        """更新BOM记录（自动清洗数据）"""
        cleaned = {
            "finished_project_no": self._clean_value(data.get("成品项目号", data.get("finished_project_no", "")), "str") or "",
            "product_name": self._clean_value(data.get("品名", data.get("product_name", "")), "str"),
            "spec": self._clean_value(data.get("规格", data.get("spec", "")), "str"),
            "retail_price": self._clean_value(data.get("零售价（元）", data.get("retail_price", 0)), "float"),
            "brand": self._clean_value(data.get("品牌", data.get("brand", "")), "str"),
            "material_project_no": self._clean_value(data.get("物料项目号", data.get("material_project_no", "")), "str") or "",
            "material_name": self._clean_value(data.get("物料名称", data.get("material_name", "")), "str"),
            "quantity": self._clean_value(data.get("数量", data.get("quantity", 0)), "float"),
            "unit": self._clean_value(data.get("计量单位", data.get("unit", "")), "str"),
            "id": bid,
        }
        c = self.conn.cursor()
        c.execute("""
            UPDATE product_bom SET finished_project_no=:finished_project_no,
            product_name=:product_name, spec=:spec, retail_price=:retail_price,
            brand=:brand, material_project_no=:material_project_no,
            material_name=:material_name, quantity=:quantity, unit=:unit
            WHERE id=:id
        """, cleaned)
        self.conn.commit()

    def delete_product_bom(self, bid):
        c = self.conn.cursor()
        c.execute("DELETE FROM product_bom WHERE id=?", (bid,))
        self.conn.commit()

    def import_product_bom(self, rows):
        """批量导入BOM数据（自动清洗数据）"""
        c = self.conn.cursor()
        for row in rows:
            cleaned = {
                "finished_project_no": self._clean_value(row.get("成品项目号", row.get("finished_project_no", "")), "str") or "",
                "product_name": self._clean_value(row.get("品名", row.get("product_name", "")), "str"),
                "spec": self._clean_value(row.get("规格", row.get("spec", "")), "str"),
                "retail_price": self._clean_value(row.get("零售价（元）", row.get("retail_price", 0)), "float"),
                "brand": self._clean_value(row.get("品牌", row.get("brand", "")), "str"),
                "material_project_no": self._clean_value(row.get("物料项目号", row.get("material_project_no", "")), "str") or "",
                "material_name": self._clean_value(row.get("物料名称", row.get("material_name", "")), "str"),
                "quantity": self._clean_value(row.get("数量", row.get("quantity", 0)), "float"),
                "unit": self._clean_value(row.get("计量单位", row.get("unit", "")), "str"),
            }
            if not cleaned["finished_project_no"] or not cleaned["material_project_no"]:
                continue
            c.execute("""
                INSERT INTO product_bom(finished_project_no, product_name, spec,
                retail_price, brand, material_project_no, material_name, quantity, unit)
                VALUES(:finished_project_no, :product_name, :spec,
                :retail_price, :brand, :material_project_no, :material_name, :quantity, :unit)
            """, cleaned)
        self.conn.commit()

    def close(self):
        self.conn.close()

    # ====== 通用Excel导出 ======
    @staticmethod
    def export_to_xlsx(filepath, sheet_name, headers, rows, col_widths=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            hdr_fill = PatternFill(start_color="C1816D", end_color="C1816D", fill_type="solid")
            hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
            hdr_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = hdr_align
                cell.border = thin_border

            data_font = Font(name="微软雅黑", size=10)
            data_align = Alignment(vertical="center")
            alt_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")

            for ri, row in enumerate(rows, 2):
                for ci, key in enumerate(headers, 1):
                    val = row.get(key, "") if isinstance(row, dict) else ""
                    cell = ws.cell(row=ri, column=ci, value=val if val is not None else "")
                    cell.font = data_font
                    cell.alignment = data_align
                    cell.border = thin_border
                    if ri % 2 == 0:
                        cell.fill = alt_fill

            if col_widths:
                for ci, w in enumerate(col_widths, 1):
                    ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
            else:
                for ci in range(1, len(headers) + 1):
                    ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 16

            wb.save(filepath)
            return True
        except Exception as e:
            raise e
